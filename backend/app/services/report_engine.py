"""Report Generation Engine — generates PDF, Markdown, and Excel reports for each platform tab."""

import io
import logging
from datetime import datetime, timezone
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)


class ReportEngine:
    """Generates reports in PDF, Markdown, and Excel formats."""

    def __init__(self):
        self.generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # ─── PDF Generation ───────────────────────────────────────────────────

    def _create_pdf_doc(self, buffer: io.BytesIO, title: str, landscape_mode: bool = False) -> SimpleDocTemplate:
        page_size = landscape(A4) if landscape_mode else A4
        return SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            title=title,
            author="MII Platform",
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

    def _pdf_header(self, title: str, subtitle: str = "") -> list:
        styles = getSampleStyleSheet()
        elements = []
        # Title
        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Title"], fontSize=20, spaceAfter=6, textColor=colors.HexColor("#1E293B")
        )
        elements.append(Paragraph(title, title_style))
        # Subtitle
        if subtitle:
            sub_style = ParagraphStyle(
                "ReportSubtitle", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#64748B")
            )
            elements.append(Paragraph(subtitle, sub_style))
        # Generated at
        meta_style = ParagraphStyle(
            "ReportMeta", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#94A3B8"), spaceAfter=20
        )
        elements.append(Paragraph(f"Generated: {self.generated_at} | Machine Identity Intelligence Platform", meta_style))
        elements.append(Spacer(1, 12))
        return elements

    def _pdf_table(self, headers: list[str], rows: list[list[str]], col_widths: list[float] | None = None) -> Table:
        data = [headers] + rows
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        return table

    # ─── Excel Generation ─────────────────────────────────────────────────

    def _create_excel_workbook(self) -> Workbook:
        wb = Workbook()
        return wb

    def _excel_header_style(self) -> dict:
        return {
            "font": Font(bold=True, color="FFFFFF", size=10),
            "fill": PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": Border(
                bottom=Side(style="thin", color="E2E8F0"),
            ),
        }

    def _excel_write_headers(self, ws, headers: list[str]):
        style = self._excel_header_style()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = style["font"]
            cell.fill = style["fill"]
            cell.alignment = style["alignment"]
            cell.border = style["border"]

    def _excel_auto_width(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

    # ─── IDENTITIES REPORT ────────────────────────────────────────────────

    def generate_identities_pdf(self, identities: list[dict]) -> bytes:
        buffer = io.BytesIO()
        doc = self._create_pdf_doc(buffer, "Identity Inventory Report", landscape_mode=True)
        elements = self._pdf_header(
            "Identity Inventory Report",
            f"{len(identities)} machine identities discovered across AWS and CI/CD"
        )

        headers = ["Name", "Type", "Source", "Risk Score", "Last Used", "Trust Count"]
        rows = []
        for i in identities:
            rows.append([
                str(i.get("name", ""))[:40],
                str(i.get("type", "")),
                str(i.get("source", "")),
                str(i.get("risk_score", 0)),
                str(i.get("last_used", "Never"))[:10],
                str(i.get("trust_count", 0)),
            ])

        elements.append(self._pdf_table(headers, rows, col_widths=[2.5 * inch, 1.2 * inch, 1 * inch, 0.8 * inch, 1 * inch, 0.8 * inch]))
        doc.build(elements)
        return buffer.getvalue()

    def generate_identities_markdown(self, identities: list[dict]) -> str:
        lines = [
            "# Identity Inventory Report",
            f"",
            f"**Generated:** {self.generated_at}",
            f"**Total Identities:** {len(identities)}",
            "",
            "---",
            "",
            "| Name | Type | Source | Risk Score | Last Used | Trust Count |",
            "|------|------|--------|-----------|-----------|-------------|",
        ]
        for i in identities:
            lines.append(
                f"| {i.get('name', '')} | {i.get('type', '')} | {i.get('source', '')} "
                f"| {i.get('risk_score', 0)} | {i.get('last_used', 'Never')} | {i.get('trust_count', 0)} |"
            )
        lines.append("")
        lines.append("---")
        lines.append(f"*Report generated by Machine Identity Intelligence Platform*")
        return "\n".join(lines)

    def generate_identities_excel(self, identities: list[dict]) -> bytes:
        wb = self._create_excel_workbook()
        ws = wb.active
        ws.title = "Identities"

        headers = ["Name", "Type", "Source", "Risk Score", "Last Used", "Trust Count"]
        self._excel_write_headers(ws, headers)

        for row_idx, i in enumerate(identities, 2):
            ws.cell(row=row_idx, column=1, value=i.get("name", ""))
            ws.cell(row=row_idx, column=2, value=i.get("type", ""))
            ws.cell(row=row_idx, column=3, value=i.get("source", ""))
            ws.cell(row=row_idx, column=4, value=i.get("risk_score", 0))
            ws.cell(row=row_idx, column=5, value=str(i.get("last_used", "Never")))
            ws.cell(row=row_idx, column=6, value=i.get("trust_count", 0))

        self._excel_auto_width(ws)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ─── FINDINGS REPORT ──────────────────────────────────────────────────

    def generate_findings_pdf(self, data: dict) -> bytes:
        buffer = io.BytesIO()
        doc = self._create_pdf_doc(buffer, "Security Findings Report", landscape_mode=True)
        findings = data.get("findings", [])
        summary = data.get("summary", {})

        elements = self._pdf_header(
            "Security Findings Report",
            f"{len(findings)} findings — {summary.get('critical', 0)} critical, "
            f"{summary.get('high', 0)} high, {summary.get('medium', 0)} medium"
        )

        headers = ["Severity", "Title", "Category", "Affected Identity", "Blast Radius", "Remediation"]
        rows = []
        for f in findings:
            rows.append([
                f.get("severity", "").upper(),
                str(f.get("title", ""))[:35],
                str(f.get("category", "")),
                str(f.get("affected_identity_name", ""))[:25],
                str(f.get("blast_radius", ""))[:30],
                str(f.get("remediation", ""))[:40],
            ])

        elements.append(self._pdf_table(headers, rows, col_widths=[0.7 * inch, 2 * inch, 1.2 * inch, 1.5 * inch, 1.8 * inch, 2.5 * inch]))
        doc.build(elements)
        return buffer.getvalue()

    def generate_findings_markdown(self, data: dict) -> str:
        findings = data.get("findings", [])
        summary = data.get("summary", {})

        lines = [
            "# Security Findings Report",
            "",
            f"**Generated:** {self.generated_at}",
            f"**Total Findings:** {len(findings)}",
            "",
            "## Summary",
            "",
            f"| Severity | Count |",
            f"|----------|-------|",
            f"| Critical | {summary.get('critical', 0)} |",
            f"| High | {summary.get('high', 0)} |",
            f"| Medium | {summary.get('medium', 0)} |",
            f"| Low | {summary.get('low', 0)} |",
            "",
            "---",
            "",
            "## Findings",
            "",
        ]
        for f in findings:
            lines.append(f"### [{f.get('severity', '').upper()}] {f.get('title', '')}")
            lines.append(f"")
            lines.append(f"- **Category:** {f.get('category', '')}")
            lines.append(f"- **Affected:** {f.get('affected_identity_name', '')}")
            lines.append(f"- **Blast Radius:** {f.get('blast_radius', '')}")
            lines.append(f"- **Description:** {f.get('description', '')}")
            lines.append(f"- **Remediation:** {f.get('remediation', '')}")
            if f.get("remediation_command"):
                lines.append(f"")
                lines.append(f"```bash")
                lines.append(f"{f['remediation_command']}")
                lines.append(f"```")
            lines.append("")

        lines.append("---")
        lines.append(f"*Report generated by Machine Identity Intelligence Platform*")
        return "\n".join(lines)

    def generate_findings_excel(self, data: dict) -> bytes:
        wb = self._create_excel_workbook()
        ws = wb.active
        ws.title = "Findings"
        findings = data.get("findings", [])

        headers = ["Severity", "Title", "Category", "Affected Identity", "Description", "Blast Radius", "Remediation", "Command"]
        self._excel_write_headers(ws, headers)

        for row_idx, f in enumerate(findings, 2):
            ws.cell(row=row_idx, column=1, value=f.get("severity", "").upper())
            ws.cell(row=row_idx, column=2, value=f.get("title", ""))
            ws.cell(row=row_idx, column=3, value=f.get("category", ""))
            ws.cell(row=row_idx, column=4, value=f.get("affected_identity_name", ""))
            ws.cell(row=row_idx, column=5, value=f.get("description", ""))
            ws.cell(row=row_idx, column=6, value=f.get("blast_radius", ""))
            ws.cell(row=row_idx, column=7, value=f.get("remediation", ""))
            ws.cell(row=row_idx, column=8, value=f.get("remediation_command", ""))

        self._excel_auto_width(ws)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ─── COMPLIANCE REPORT ────────────────────────────────────────────────

    def generate_compliance_pdf(self, data: dict) -> bytes:
        buffer = io.BytesIO()
        doc = self._create_pdf_doc(buffer, "Compliance Report")
        checks = data.get("checks", [])
        score = data.get("compliance_score", 0)

        elements = self._pdf_header(
            "Compliance Report",
            f"Compliance Score: {score}% — {data.get('passing', 0)} passing, {data.get('failing', 0)} failing"
        )

        headers = ["ID", "Policy", "Status", "Severity", "Passing", "Failing", "Recommendation"]
        rows = []
        for c in checks:
            rows.append([
                c.get("id", ""),
                str(c.get("policy_name", ""))[:30],
                c.get("status", "").upper(),
                c.get("severity", ""),
                str(c.get("passing", 0)),
                str(c.get("failing", 0)),
                str(c.get("recommendation", ""))[:40],
            ])

        elements.append(self._pdf_table(headers, rows))
        doc.build(elements)
        return buffer.getvalue()

    def generate_compliance_markdown(self, data: dict) -> str:
        checks = data.get("checks", [])
        score = data.get("compliance_score", 0)

        lines = [
            "# Compliance Report",
            "",
            f"**Generated:** {self.generated_at}",
            f"**Compliance Score:** {score}%",
            f"**Passing:** {data.get('passing', 0)} | **Failing:** {data.get('failing', 0)} | **Warnings:** {data.get('warnings', 0)}",
            "",
            "---",
            "",
            "## Policy Checks",
            "",
            "| ID | Policy | Status | Severity | Passing | Failing |",
            "|----|--------|--------|----------|---------|---------|",
        ]
        for c in checks:
            status_icon = "PASS" if c.get("status") == "pass" else "FAIL"
            lines.append(
                f"| {c.get('id', '')} | {c.get('policy_name', '')} | {status_icon} "
                f"| {c.get('severity', '')} | {c.get('passing', 0)} | {c.get('failing', 0)} |"
            )

        lines.append("")
        lines.append("## Details")
        lines.append("")
        for c in checks:
            if c.get("status") == "fail":
                lines.append(f"### {c.get('id', '')} — {c.get('policy_name', '')} (FAIL)")
                lines.append(f"")
                lines.append(f"- **Description:** {c.get('description', '')}")
                lines.append(f"- **Recommendation:** {c.get('recommendation', '')}")
                failing_ids = c.get("failing_identities", [])
                if failing_ids:
                    lines.append(f"- **Failing Identities:**")
                    for fi in failing_ids[:10]:
                        lines.append(f"  - {fi.get('name', '')} — {fi.get('reason', '')}")
                lines.append("")

        lines.append("---")
        lines.append(f"*Report generated by Machine Identity Intelligence Platform*")
        return "\n".join(lines)

    def generate_compliance_excel(self, data: dict) -> bytes:
        wb = self._create_excel_workbook()
        ws = wb.active
        ws.title = "Compliance"
        checks = data.get("checks", [])

        headers = ["ID", "Policy Name", "Status", "Severity", "Description", "Total Checked", "Passing", "Failing", "Recommendation"]
        self._excel_write_headers(ws, headers)

        for row_idx, c in enumerate(checks, 2):
            ws.cell(row=row_idx, column=1, value=c.get("id", ""))
            ws.cell(row=row_idx, column=2, value=c.get("policy_name", ""))
            ws.cell(row=row_idx, column=3, value=c.get("status", "").upper())
            ws.cell(row=row_idx, column=4, value=c.get("severity", ""))
            ws.cell(row=row_idx, column=5, value=c.get("description", ""))
            ws.cell(row=row_idx, column=6, value=c.get("total_checked", 0))
            ws.cell(row=row_idx, column=7, value=c.get("passing", 0))
            ws.cell(row=row_idx, column=8, value=c.get("failing", 0))
            ws.cell(row=row_idx, column=9, value=c.get("recommendation", ""))

        # Add failing identities sheet
        if any(c.get("failing_identities") for c in checks):
            ws2 = wb.create_sheet("Failing Identities")
            headers2 = ["Check ID", "Policy", "Identity Name", "Reason"]
            self._excel_write_headers(ws2, headers2)
            row = 2
            for c in checks:
                for fi in c.get("failing_identities", []):
                    ws2.cell(row=row, column=1, value=c.get("id", ""))
                    ws2.cell(row=row, column=2, value=c.get("policy_name", ""))
                    ws2.cell(row=row, column=3, value=fi.get("name", ""))
                    ws2.cell(row=row, column=4, value=fi.get("reason", ""))
                    row += 1
            self._excel_auto_width(ws2)

        self._excel_auto_width(ws)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ─── TRUST DEBT REPORT ────────────────────────────────────────────────

    def generate_trust_debt_pdf(self, data: dict) -> bytes:
        buffer = io.BytesIO()
        doc = self._create_pdf_doc(buffer, "Trust Debt Report")
        debt_items = data.get("debt_items", [])
        grade = data.get("debt_grade", "?")
        total_score = data.get("total_debt_score", 0)

        elements = self._pdf_header(
            "Trust Debt Report",
            f"Grade: {grade} | Debt Score: {total_score} | {len(debt_items)} debt items identified"
        )

        headers = ["Category", "Identity", "Debt Points", "Description", "Remediation"]
        rows = []
        for item in debt_items:
            rows.append([
                str(item.get("category", "")),
                str(item.get("identity_name", ""))[:25],
                str(item.get("debt_points", 0)),
                str(item.get("description", ""))[:35],
                str(item.get("remediation", ""))[:35],
            ])

        elements.append(self._pdf_table(headers, rows))
        doc.build(elements)
        return buffer.getvalue()

    def generate_trust_debt_markdown(self, data: dict) -> str:
        debt_items = data.get("debt_items", [])
        grade = data.get("debt_grade", "?")
        total_score = data.get("total_debt_score", 0)

        lines = [
            "# Trust Debt Report",
            "",
            f"**Generated:** {self.generated_at}",
            f"**Grade:** {grade}",
            f"**Total Debt Score:** {total_score}",
            f"**Debt Items:** {len(debt_items)}",
            "",
            "---",
            "",
            "## Debt Items",
            "",
            "| Category | Identity | Debt Points | Description |",
            "|----------|----------|-------------|-------------|",
        ]
        for item in debt_items:
            lines.append(
                f"| {item.get('category', '')} | {item.get('identity_name', '')} "
                f"| {item.get('debt_points', 0)} | {item.get('description', '')} |"
            )

        lines.append("")

        # Category breakdown
        categories: dict[str, int] = {}
        for item in debt_items:
            cat = item.get("category", "Unknown")
            categories[cat] = categories.get(cat, 0) + item.get("debt_points", 0)

        if categories:
            lines.append("## Category Breakdown")
            lines.append("")
            lines.append("| Category | Total Debt Points |")
            lines.append("|----------|-------------------|")
            for cat, points in sorted(categories.items(), key=lambda x: x[1], reverse=True):
                lines.append(f"| {cat} | {points} |")
            lines.append("")

        lines.append("---")
        lines.append(f"*Report generated by Machine Identity Intelligence Platform*")
        return "\n".join(lines)

    def generate_trust_debt_excel(self, data: dict) -> bytes:
        wb = self._create_excel_workbook()
        ws = wb.active
        ws.title = "Trust Debt"
        debt_items = data.get("debt_items", [])

        headers = ["Category", "Identity", "Debt Points", "Description", "Remediation"]
        self._excel_write_headers(ws, headers)

        for row_idx, item in enumerate(debt_items, 2):
            ws.cell(row=row_idx, column=1, value=item.get("category", ""))
            ws.cell(row=row_idx, column=2, value=item.get("identity_name", ""))
            ws.cell(row=row_idx, column=3, value=item.get("debt_points", 0))
            ws.cell(row=row_idx, column=4, value=item.get("description", ""))
            ws.cell(row=row_idx, column=5, value=item.get("remediation", ""))

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.cell(row=1, column=1, value="Metric")
        ws2.cell(row=1, column=2, value="Value")
        ws2.cell(row=2, column=1, value="Grade")
        ws2.cell(row=2, column=2, value=data.get("debt_grade", "?"))
        ws2.cell(row=3, column=1, value="Total Debt Score")
        ws2.cell(row=3, column=2, value=data.get("total_debt_score", 0))
        ws2.cell(row=4, column=1, value="Total Items")
        ws2.cell(row=4, column=2, value=len(debt_items))

        self._excel_auto_width(ws)
        self._excel_auto_width(ws2)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ─── BLAST PATH REPORT ────────────────────────────────────────────────

    def generate_blast_path_pdf(self, data: dict) -> bytes:
        buffer = io.BytesIO()
        doc = self._create_pdf_doc(buffer, "Blast Path Simulation Report")
        steps = data.get("steps", [])
        source = data.get("source_identity", "Unknown")
        severity = data.get("overall_severity", "Unknown")

        elements = self._pdf_header(
            "Blast Path Simulation Report",
            f"Source: {source} | Overall Severity: {severity} | {len(steps)} attack steps"
        )

        headers = ["Step", "Identity", "Action", "Risk", "Description"]
        rows = []
        for idx, step in enumerate(steps, 1):
            rows.append([
                str(idx),
                str(step.get("identity_name", step.get("target", "")))[:25],
                str(step.get("action", step.get("relationship", "")))[:20],
                str(step.get("risk_level", step.get("severity", "")))[:10],
                str(step.get("description", step.get("explanation", "")))[:40],
            ])

        elements.append(self._pdf_table(headers, rows))
        doc.build(elements)
        return buffer.getvalue()

    def generate_blast_path_markdown(self, data: dict) -> str:
        steps = data.get("steps", [])
        source = data.get("source_identity", "Unknown")
        severity = data.get("overall_severity", "Unknown")

        lines = [
            "# Blast Path Simulation Report",
            "",
            f"**Generated:** {self.generated_at}",
            f"**Source Identity:** {source}",
            f"**Overall Severity:** {severity}",
            f"**Attack Steps:** {len(steps)}",
            "",
            "---",
            "",
            "## Attack Chain",
            "",
        ]
        for idx, step in enumerate(steps, 1):
            identity = step.get("identity_name", step.get("target", ""))
            action = step.get("action", step.get("relationship", ""))
            risk = step.get("risk_level", step.get("severity", ""))
            desc = step.get("description", step.get("explanation", ""))
            lines.append(f"### Step {idx}: {identity}")
            lines.append(f"")
            lines.append(f"- **Action:** {action}")
            lines.append(f"- **Risk Level:** {risk}")
            lines.append(f"- **Description:** {desc}")
            lines.append("")

        lines.append("---")
        lines.append(f"*Report generated by Machine Identity Intelligence Platform*")
        return "\n".join(lines)

    def generate_blast_path_excel(self, data: dict) -> bytes:
        wb = self._create_excel_workbook()
        ws = wb.active
        ws.title = "Blast Path"
        steps = data.get("steps", [])

        headers = ["Step", "Identity", "Action", "Risk Level", "Description"]
        self._excel_write_headers(ws, headers)

        for row_idx, step in enumerate(steps, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=step.get("identity_name", step.get("target", "")))
            ws.cell(row=row_idx, column=3, value=step.get("action", step.get("relationship", "")))
            ws.cell(row=row_idx, column=4, value=step.get("risk_level", step.get("severity", "")))
            ws.cell(row=row_idx, column=5, value=step.get("description", step.get("explanation", "")))

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.cell(row=1, column=1, value="Metric")
        ws2.cell(row=1, column=2, value="Value")
        ws2.cell(row=2, column=1, value="Source Identity")
        ws2.cell(row=2, column=2, value=data.get("source_identity", "Unknown"))
        ws2.cell(row=3, column=1, value="Overall Severity")
        ws2.cell(row=3, column=2, value=data.get("overall_severity", "Unknown"))
        ws2.cell(row=4, column=1, value="Total Steps")
        ws2.cell(row=4, column=2, value=len(steps))

        self._excel_auto_width(ws)
        self._excel_auto_width(ws2)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
